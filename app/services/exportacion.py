from io import BytesIO
from datetime import date, timedelta


from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter


def restar_meses(fecha, meses):
    """
    Resta meses a una fecha sin utilizar
    librerías externas.
    """

    año = fecha.year

    mes = fecha.month - meses


    while mes <= 0:

        mes += 12

        año -= 1


    return date(
        año,
        mes,
        1
    )


def generar_excel_asistencia(
    registros,
    periodo="todos",
    mes=None
):

    """
    Genera el Excel de asistencia.

    periodo:

        "mes"
            Mes específico.
            Requiere mes="YYYY-MM"

        "3_meses"
            Mes actual + 2 meses anteriores.

        "6_meses"
            Mes actual + 5 meses anteriores.

        "todos"
            Todos los registros.

    mes:

        Formato YYYY-MM.

        Ejemplo:

        2026-10
    """


    # ==========================================
    # CREAR LIBRO
    # ==========================================

    libro = Workbook()

    hoja = libro.active

    hoja.title = "Asistencia"


    # ==========================================
    # FECHA ACTUAL
    # ==========================================

    hoy = date.today()


    # ==========================================
    # RANGO DE FECHAS
    # ==========================================

    fecha_inicio = None

    fecha_fin = None


    # =====================================================
    # MES ESPECÍFICO
    # =====================================================

    if periodo == "mes":

        if not mes:

            raise ValueError(
                "Debes seleccionar un mes."
            )


        try:

            partes = mes.split("-")

            if len(partes) != 2:

                raise ValueError


            año = int(
                partes[0]
            )

            numero_mes = int(
                partes[1]
            )


            if numero_mes < 1 or numero_mes > 12:

                raise ValueError


            fecha_inicio = date(
                año,
                numero_mes,
                1
            )


            # Primer día del siguiente mes

            if numero_mes == 12:

                siguiente_mes = date(
                    año + 1,
                    1,
                    1
                )

            else:

                siguiente_mes = date(
                    año,
                    numero_mes + 1,
                    1
                )


            # Último día del mes seleccionado

            fecha_fin = siguiente_mes - timedelta(days=1)



        except (ValueError, TypeError):

            raise ValueError(
                "El mes debe tener el formato YYYY-MM."
            )


    # =====================================================
    # ÚLTIMOS 3 MESES
    # =====================================================

    elif periodo == "3_meses":

        fecha_inicio = restar_meses(
            hoy,
            2
        )

        fecha_fin = hoy


    # =====================================================
    # ÚLTIMOS 6 MESES
    # =====================================================

    elif periodo == "6_meses":

        fecha_inicio = restar_meses(
            hoy,
            5
        )

        fecha_fin = hoy


    # =====================================================
    # TODOS
    # =====================================================

    elif periodo == "todos":

        fecha_inicio = None

        fecha_fin = None


    else:

        raise ValueError(
            "Periodo inválido."
        )


    # ==========================================
    # ENCABEZADOS
    # ==========================================

    encabezados = [

        "ID Jornada",

        "Empleado",

        "Correo",

        "Fecha",

        "Entrada",

        "Break mañana",

        "Lunch",

        "Break tarde",

        "Salida"

    ]


    hoja.append(
        encabezados
    )


    # ==========================================
    # ESTILO ENCABEZADOS
    # ==========================================

    color_azul = "0D6EFD"


    fondo = PatternFill(

        fill_type="solid",

        fgColor=color_azul

    )


    borde = Border(

        bottom=Side(

            style="thin",

            color="FFFFFF"

        )

    )


    for celda in hoja[1]:

        celda.font = Font(

            bold=True,

            color="FFFFFF"

        )


        celda.fill = fondo


        celda.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )


        celda.border = borde


    # ==========================================
    # FORMATO HORA
    # ==========================================

    def formato_hora(hora):

        if hora is None:

            return ""


        return hora.strftime(
            "%H:%M:%S"
        )


    # ==========================================
    # FORMATO DESCANSO
    # ==========================================

    def formato_descanso(descanso):

        if descanso is None:

            return ""


        inicio = formato_hora(
            descanso.inicio
        )


        if descanso.fin:

            fin = formato_hora(
                descanso.fin
            )

        else:

            fin = "En curso"


        return (
            f"{inicio} - {fin}"
        )


    # ==========================================
    # RECORRER JORNADAS
    # ==========================================

    for jornada in registros:

        usuario = jornada.usuario


        # --------------------------------------
        # USUARIO NO EXISTE
        # --------------------------------------

        if usuario is None:

            continue


        # --------------------------------------
        # NO EXPORTAR ADMIN
        # --------------------------------------

        if (
            str(usuario.rol)
            .strip()
            .lower()
            == "admin"
        ):

            continue


        # ======================================
        # FILTRO POR FECHA
        # ======================================

        if periodo != "todos":

            if not jornada.fecha:

                continue


            fecha_jornada = jornada.fecha


            # Si fecha es datetime
            if hasattr(
                fecha_jornada,
                "date"
            ):

                fecha_jornada = (
                    fecha_jornada.date()
                )


            # ----------------------------------
            # COMPARAR FECHA
            # ----------------------------------

            if (
                fecha_jornada < fecha_inicio
                or fecha_jornada > fecha_fin
            ):

                continue


        # ======================================
        # DESCANSOS
        # ======================================

        break_manana = None

        lunch = None

        break_tarde = None


        for descanso in jornada.descansos:

            if (
                descanso.tipo
                == "break_manana"
            ):

                break_manana = descanso


            elif (
                descanso.tipo
                == "lunch"
            ):

                lunch = descanso


            elif (
                descanso.tipo
                == "break_tarde"
            ):

                break_tarde = descanso


        # ======================================
        # AGREGAR FILA
        # ======================================

        hoja.append([

            jornada.id,

            usuario.nombre,

            usuario.correo,

            jornada.fecha.strftime(
                "%Y-%m-%d"
            )
            if jornada.fecha
            else "",

            formato_hora(
                jornada.entrada
            ),

            formato_descanso(
                break_manana
            ),

            formato_descanso(
                lunch
            ),

            formato_descanso(
                break_tarde
            ),

            formato_hora(
                jornada.salida
            )

        ])


    # ==========================================
    # AJUSTAR COLUMNAS
    # ==========================================

    anchos = {

        1: 14,

        2: 25,

        3: 35,

        4: 15,

        5: 15,

        6: 22,

        7: 22,

        8: 22,

        9: 15

    }


    for numero_columna, ancho in anchos.items():

        letra = get_column_letter(
            numero_columna
        )


        hoja.column_dimensions[
            letra
        ].width = ancho


    # ==========================================
    # CENTRAR COLUMNAS
    # ==========================================

    for fila in hoja.iter_rows(
        min_row=2
    ):

        for indice in [

            0,

            3,

            4,

            5,

            6,

            7,

            8

        ]:

            fila[indice].alignment = Alignment(

                horizontal="center"

            )


    # ==========================================
    # CONGELAR ENCABEZADO
    # ==========================================

    hoja.freeze_panes = "A2"


    # ==========================================
    # FILTRO EXCEL
    # ==========================================

    hoja.auto_filter.ref = hoja.dimensions


    # ==========================================
    # ALTURA ENCABEZADO
    # ==========================================

    hoja.row_dimensions[1].height = 25


    # ==========================================
    # CREAR ARCHIVO
    # ==========================================

    archivo = BytesIO()


    libro.save(
        archivo
    )


    archivo.seek(0)


    return archivo
